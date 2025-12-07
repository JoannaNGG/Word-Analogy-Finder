import gensim.downloader as api
import numpy as np
from openpyxl import Workbook, load_workbook
import os

print("\nLoading the Word2Vec Model, please wait...")
model = api.load("word2vec-google-news-300")
print("Model Loaded.\n")

excel_file = "pairsim_results.xlsx"

#Create the excel file
if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["A", "B", "C", "D", "Similarity", "Same_Relation"])
    wb.save(excel_file)

def relation_similarity(A, B, C, D):
    vec1 = model[B] - model[A]
    vec2 = model[D] - model[C]
    similarity = np.dot(vec1, vec2) / (np.linalg.norm(vec1) * np.linalg.norm(vec2))
    return similarity

#Looping for multiple analogies
while True:
    print("\nEnter your first word pair (A:B): ")
    pair1 = input().strip().lower()

    print("Enter your second word pair (C:D): ")
    pair2 = input().strip().lower()

    try:
        A, B = [w.strip() for w in pair1.split(":")]
        C, D = [w.strip() for w in pair2.split(":")]
    except:
        print("Invalid format.")
        continue

    #Compute similarity
    relation_sim = relation_similarity(A, B, C, D)
    same_relation = "YES" if relation_sim >= 0.8 else "NO"

    print(f"\nRelationship similarity between ({A}:{B}) and ({C}:{D}) = {relation_sim:.3f}")
    print("The pairs have approximately the same relationship." 
          if same_relation=="YES"
          else "The pairs do NOT have the same relationship.")

    #Save results to excel sheet
    wb = load_workbook(excel_file)
    ws = wb["Results"]

    ws.append([A, B, C, D, round(relation_sim, 3), same_relation])
    wb.save(excel_file)

    print(f"Saved results to: {excel_file}")

    #Asking user whether to continue
    again = input("\nTest another analogy? (y/n): ").strip().lower()
    if again != "y":
        print("\nSession ended. All results saved.")
        break