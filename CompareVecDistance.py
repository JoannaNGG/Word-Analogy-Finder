import gensim.downloader as api
from openpyxl import Workbook, load_workbook
import os

print("\nLoading the Word2Vec Model, please wait...")
model = api.load("word2vec-google-news-300")
print("Model Loaded.\n")

#create and save to exccel file
excel_file = "analogy_vec_distance_results.xlsx"

if not os.path.exists(excel_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Analogies"
    ws.append(["A", "B", "C", "Result", "TopScore", "Sim_to_Male", "Sim_to_Female"])
    wb.save(excel_file)

#Looping for multiple analogies
while True:
    analogy = input("\nEnter an analogy in format 'A:B::C:?'\n").strip()

    try:
        left, right = analogy.split("::")
        A, B = left.split(":")
        C, _ = right.split(":")
        A, B, C = A.strip().lower(), B.strip().lower(), C.strip().lower()
    except:
        print("Invalid format. Use A:B::C:?")
        continue

    #Compute B - A + C
    result = model.most_similar(
        positive=[B.strip(), C.strip()],
        negative=[A.strip()],
        topn=5
    )

    roundResult = [(word, round(score, 3)) for word, score in result]
    print(roundResult)

    #Compare top 1 result with "male/female"
    top_word, top_score = result[0]
    top_score = round(top_score, 3)

    sim_male = round(model.similarity(top_word, "male"), 3)
    sim_female = round(model.similarity(top_word, "female"), 3)

    print(f"\nSimilarity of '{top_word}' to 'male': {sim_male:.3f}")
    print(f"Similarity of '{top_word}' to 'female': {sim_female:.3f}")
    
    #Save results to excel sheet
    wb = load_workbook(excel_file)
    ws = wb["Analogies"]

    ws.append([
        A, B, C,
        top_word,
        top_score,
        sim_male,
        sim_female
    ])

    wb.save(excel_file)
    print(f"Saved results to: {excel_file}")

    #Asking user whether to continue
    again = input("\nTest another analogy? (y/n): ").strip().lower()
    if again != "y":
        print("\nSession ended. All results saved.")
        break